# -*- coding: utf-8 -*-
"""
Created on Sat Mar  4 11:02:52 2017

Description: Save data to xls

Scope: Modelling circular economy policies in EEIOA


@author: Franco Donati
@institution: Leiden University CML
"""
import os
import datetime
import pandas as pd
import pickle as pk
from shutil import copyfile
from pandas import DataFrame as df
from openpyxl import load_workbook
now = datetime.datetime.now()

def save_outputs(results, directory, specs, scen_no, data=None):
    """
    It saves results into a previously specified directory

    Parameters
    ----------

    results : pandas.DataFrame
        a dataframe containing only selected results

    directory : str
        the general location specified for your scenarios.xlsx file

    specs : dict
        a dictionary containing general info about the project (e.g. author, institution, etc)

    scen_no : int or None
        an integer specifying the scenario number or enter None to save all results

    data : dict of pd.DataFrames
        a completely new dataset to be pickled. Default value is None otherwise
        pass the dataset

    Output
    ------
    scenarios.xlsx : excel file
        scenario settings excel file used for the analysis in the same output directory with the results

    info_and_results.xlsx : excel file
        excel file containing general info about project plus the results from the analysis

    data.pkl : pickle file
        new modified IO dataset in pickle format

    """
    date = "_".join([str(now.year), str(now.month), str(now.day)])
    time = "_".join([str(now.hour), str(now.minute)])
    directory_or = directory
    directory_out = os.path.join(directory, "output_" + date)

    if scen_no in ["baseline", 0, "base"]:
        scen_no = "baseline"
        path = os.path.join(directory_out, "baseline", time)

    elif type(scen_no) is int:
        if scen_no > 0:
            scen_no = "scenario_" + str(scen_no)
            specs["scenario_number"] = scen_no
            path = os.path.join(directory_out, "_".join([str(scen_no)]), time)

    elif scen_no is None:
        scen_no = "all_results"
        specs["scenario_number"] = scen_no
        path = os.path.join(directory_out, "all_results", time)

    if not os.path.exists(path):
        os.makedirs(path)

    specs = add_date_to_gen_specs(specs)

    copyfile(os.path.join(directory_or, "scenarios.xlsx"), os.path.join(path, "scenarios.xlsx"))

    specs.to_excel(os.path.join(path, "info_and_results.xlsx"), sheet_name="Info", startrow=1, startcol=1)
    with pd.ExcelWriter(os.path.join(path, "info_and_results.xlsx"), engine="openpyxl") as writer:
        writer.book = load_workbook(os.path.join(path, "info_and_results.xlsx"))
        results.to_excel(writer, "Results", startrow=1, startcol=1, merge_cells=False)

    if data is not None:
        w = open(path + "/"+ "data.pkl", "wb")
        pk.dump(data, w, 2)

        w.close()

def add_date_to_gen_specs(specs):
    """
    Adds timemark to general specifications e.g. authors, institution etc
    """
    year, month, day = str(now.date()).split("-")

    timemark = "/".join([day, month, year])

    specs["timemark"] = timemark

    specs = df([specs]).T
    specs.columns = ["General_info"]

    return specs
